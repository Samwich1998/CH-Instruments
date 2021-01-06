
"""
Need to Install on the Anaconda Prompt:
    $ conda install openpyxl
    $ conda install seaborn
"""
import csv
import os
import openpyxl as xl
import matplotlib.pyplot as plt
import math
import re


# -------------------------- User Can Edit ----------------------------------#

use_All_CSV_Files = True # If False, Populate the CV_CSV_Data_List Yourself
data_Directory = "../NASA Project Cortisol/Prussian Blue/12-24-2020 Precursor Analysis/" # The Folder with the CSV Files

# Specify Figure Asthetics
numSubplotWidth = 4
figWidth = 25
figHeight = 13

#numSubplotWidth = 6
#figWidth = 39*2
#figHeight = 17

# ---------------------------------------------------------------------------#

# -------------------------- Find File Names --------------------------------#

if use_All_CSV_Files:
    CV_CSV_Data_List = []
    for file in os.listdir(data_Directory):
        if file.endswith(".csv"):
            CV_CSV_Data_List.append(file)
else:
    CV_CSV_Data_List = [
         'CobaltII Sulfate + IronIII Hexacyanide 0.01M HCl + Ni.csv',
         ]
    
    # Check to see if the Inputed CSV Files Exist
    for CV_CSV_Data in CV_CSV_Data_List:    
        if not os.path.isfile(data_Directory + CV_CSV_Data):
            print("The File ", data_Directory + CV_CSV_Data," Mentioned Does NOT Exist")
            exit()

# Create Output Folder
outputData = data_Directory +  "Peak_Current_Plots/"
os.system("mkdir '" + outputData + "'")

# ---------------------------------------------------------------------------#

def addPeakCurrent(addingDict, addingKey, addingValue):
    # see if the Key is in the Dictionary
    currentValues = addingDict.get(addingKey, [])
    # If Not, Add it
    if currentValues == []:
        addingDict[addingKey] = []
    # Add the Value to the Dictionary
    addingDict[addingKey].append(addingValue)
    return addingDict

# -------------------- Extract and Plot the Ip Data -------------------------#

# For Eaqch CSV File, Extract the Important Data and Plot
fig, ax = plt.subplots(math.ceil(len(CV_CSV_Data_List)/numSubplotWidth), numSubplotWidth, sharey=False, sharex = False, figsize=(figWidth,figHeight))
errorFigNum = 0
for figNum, CV_CSV_Data in enumerate(CV_CSV_Data_List):    
    figNum -= errorFigNum
    
    # ----------------- Convert Data to Excel Format ------------------------#
    
    # Rename the File with an Excel Extension
    base = os.path.splitext(CV_CSV_Data)[0]
    excel_file = data_Directory + base + ".xlsx"
    # If the File is Not Already Converted: Convert
    if not os.path.isfile(excel_file):
        # Make Excel WorkBook
        wb = xl.Workbook()
        ws = wb.active
        # Write to Excel WorkBook
        with open(data_Directory + CV_CSV_Data) as f:
            reader = csv.reader(f, delimiter=',')
            for row in reader:
                ws.append(row)
        # Save as New Excel File
        wb.save(excel_file)
    else:
        print("You already renamed the '.csv' to '.xlsx'")
    
    # Load Data from New Excel File
    WB = xl.load_workbook(excel_file) 
    WB_worksheets = WB.worksheets
    Main = WB_worksheets[0]
        
    # -----------------------------------------------------------------------#
    
    # ----------------------- Extract Run Info ------------------------------#
    
    # Set Initial Variables from last Run to Zero
    scanRate = None; sampleInterval = None; highVolt = None; lowVolt = None; dataIndex = None
    # Loop Through the Info Section and Extract the Needxed Run Info from Excel
    for cell in Main['A']:
        # Get Cell Value
        cellVal = cell.value
        if cellVal == None:
            continue
        
        # Find the Scan Rate (Volts/Second)
        # Find the Sample Interval (Voltage Different Between Points)
        if cellVal.startswith("Sample Interval (V) = "):
            sampleInterval = float(cellVal.split(" = ")[-1])
        # Find the Highest Voltage
        elif cellVal.startswith("High E (V) = "):
            highVolt = float(cellVal.split(" = ")[-1])
        # Find the Lowest Voltage
        elif cellVal.startswith("Low E (V) = "):
            lowVolt = float(cellVal.split(" = ")[-1])
        elif cellVal == "Segment 1:":
            startSegment = cell.row
        elif cellVal == "Potential/V":
            dataIndex = cell.row + 2
            break
    # Find the X Axis Width
    xRange = (highVolt - lowVolt)*2
    # Find Point/Scan
    pointsPerScan = int(xRange/sampleInterval)
    # Total Frames (Will Round Down to Remove Incomplete Scans); Frame = Cycle = 2 Segments
    totalFrames = math.floor((Main.max_row - dataIndex + 1)/pointsPerScan)
    numberOfSegments = totalFrames*2
        
    # -----------------------------------------------------------------------#
    
    # ------------------- Find Ip Data and Store Info -----------------------#
    
    peakCurrent = {"peakCurrentForward":{}, "peakCurrentReverse":{}}
    cycleNumber = {"cycleNumberForward":{}, "cycleNumberReverse":{}}
    cycleNum = 0
    for rowA in Main.iter_rows(min_col=1, min_row=startSegment, max_col=1, max_row=dataIndex - 4):
        cellVal = rowA[0].value
        if cellVal == None:
            continue
        
        # Find the Current Segment
        if cellVal.startswith("Segment "):
            segment = float(cellVal[:-1].split("Segment ")[-1])
            peakNum = 1
            # It is a New Cycle Everytime we Scan Forwards
            if segment%2 == 1:
                cycleNum += 1
            # Stop if Next is Incomplete Segment; Else COntinue Looping
            if segment == numberOfSegments:
                break
            continue
        # find the Peak Current in the Segment
        elif cellVal.startswith("ip = "):
            Ip = [float(x.split(":")[0]) for x in re.findall("-?\d+.?\d*(?:[Ee]-\d+)?", cellVal)][0]
            # If it is an Odd Segment, it is the Forwards Peak
            if segment%2 == 1:
                peakCurrent["peakCurrentForward"] = addPeakCurrent(peakCurrent["peakCurrentForward"], peakNum, Ip)
                cycleNumber["cycleNumberForward"] = addPeakCurrent(cycleNumber["cycleNumberForward"], peakNum, cycleNum)
            else:
                peakCurrent["peakCurrentReverse"] = addPeakCurrent(peakCurrent["peakCurrentReverse"], peakNum, Ip)
                cycleNumber["cycleNumberReverse"] = addPeakCurrent(cycleNumber["cycleNumberReverse"], peakNum, cycleNum)
            peakNum += 1
    
    # ---------------------------------------------------------------------- #
  
    # ----------------------- Plot and Save the Data ----------------------- #
    
    # Decide Whether to Save the Plot
    if peakCurrent["peakCurrentForward"] == {} and peakCurrent["peakCurrentReverse"] == {}:
        print("No Peak Currents Found in: ", base)
        errorFigNum += 0
        # still plotting for some reason at end
        continue
    
    # Keep Running Subplots Order
    if numSubplotWidth == 1 and len(CV_CSV_Data_List) == 1:
        currentAxes = ax
    elif numSubplotWidth == 1:
        currentAxes = ax[figNum]
    elif numSubplotWidth == len(CV_CSV_Data_List):
        currentAxes = ax[figNum]
    elif numSubplotWidth > 1:
        currentAxes = ax[figNum//numSubplotWidth][figNum%numSubplotWidth]
    else:
        print("numSubplotWidth CANNOT be < 1. Currently it is: ", numSubplotWidth)
        exit
    
    # Plot Forwards Data
    fig1 = plt.figure(2+figNum)
    legendList = []
    
    for peakNum in peakCurrent["peakCurrentForward"].keys():
        cycleNum = cycleNumber["cycleNumberForward"][peakNum]
        peak = peakCurrent["peakCurrentForward"][peakNum]
        plt.plot(cycleNum, peak, "-o", label="Forwards Peak Number " + str(peakNum))
        currentAxes.plot(cycleNum, peak, "-o", label="Forwards Peak Number " + str(peakNum))
        legendList.append("Forwards Peak Number " + str(peakNum))
    # Plot Reverse Data
    for peakNum in peakCurrent["peakCurrentReverse"].keys():
        cycleNum = cycleNumber["cycleNumberReverse"][peakNum]
        peak = peakCurrent["peakCurrentReverse"][peakNum]
        plt.plot(cycleNum, peak, "-o", label="Reverse Peak Number " + str(peakNum))
        currentAxes.plot(cycleNum, peak, "-o", label="Reverse Peak Number " + str(peakNum))
        legendList.append("Reverse Peak Number " + str(peakNum))
    
    # Plot and Save Single Data
    plt.title("Peak Current Over CV Scan")
    plt.xlabel("Cycle Number")
    plt.ylabel("Peak Current (Amps)")
    plt.xlim(0, totalFrames)
    plt.legend(legendList, loc="best") #bbox_to_anchor=(1.05, 1))
    fig1.tight_layout(pad=3.0)
    plt.savefig(outputData + base + ".png", dpi=300)
    
    # set Title and Labels of Subplot
    currentAxes.set_xlabel("Cycle Number")
    currentAxes.set_ylabel("Peak Current (Amps)")
    currentAxes.set_xlim(0, totalFrames)
    currentAxes.set_title(base)
    currentAxes.legend() 
    currentAxes.plot()
    
    # ---------------------------------------------------------------------- #

# --------------------- Plot and Save the Data ------------------------------#

fig.tight_layout(pad=2.0)
fig.savefig(outputData + "subplots.png", dpi=300)
plt.title("Peak Current Over CV Scan") # Need this Line as we Change the Title When we Save Subplots
plt.show() # Must be the Last Line

# -------------------------------------------------------------------------- #
        
        

    