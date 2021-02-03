
"""
Need to Install Once on the Anaconda Prompt or Python Terminal:
    $ conda install openpyxl
    % conda install ffmpeg
"""

# ------------------------ Imported Packages --------------------------------#

# Import Packages
import csv
import os
import openpyxl as xl
import matplotlib.pyplot as plt
import matplotlib.animation as manimation
import math
import re
import statistics as stat

# ---------------------------------------------------------------------------#

# -------------------------- User Can Edit ----------------------------------#

# Specify Directory/Folder with All the Data (CSV Files Exported from CHI)
data_Directory = "C:/Users/weiga/Desktop/Sam/NASA Project Cortisol/Prussian Blue/2021/01-28-2021 PBS Sep Added Day 2/" # The Folder with the CSV Files

# Edit What is Plotted
skipIfDataAlreadyExcel = True  # Skip Over Data if CSV->Excel has Already Been Done (No Graphs!)
use_All_CSV_Files = True        # If False, Populate the CV_CSV_Data_List Yourself (Choose Your Files Below)
showPeakCurrent = True          # Display Real-Time Peak Current Data on Right (ONLY IF Peak Current Exists)
seePastCVData = True            # See All CSV Frames in the Background (with 10% opacity)
showFullInfo = True             # Plot Peak Potential and See Coefficient of VariationList Plot for peak Current

# Edit Data
numInitCyclesToSkip = 1         # Number of Beginning Cycles to Skip (In the First Few Cycles the Electrode is Adapting).
peakError = 0.04                # deltaV (Potential) Difference that Defines a New Peak (For Peak Labeling)

# Specify Figure Asthetics
peakCurrentRightColorOrder = {
    "peakCurrentForward|1": "tab:red",
    "peakCurrentForward|2": "tab:purple",
    "peakCurrentForward|3": "tab:orange",
    "peakCurrentForward|4": "tab:pink",
    "peakCurrentReverse|1": "tab:brown",
    "peakCurrentReverse|2": "tab:green",
    "peakCurrentReverse|3": "tab:gray",
    "peakCurrentReverse|4": "tab:cyan",
    }

# If use_All_CSV_Files is False, These Files Will be Used
CV_CSV_Data_List = [
     'Cold CuNiHCF PBS EthOH-H20 Full Run.csv',
]

# ---------------------------------------------------------------------------#

# -------------------------- Find File Names --------------------------------#

if use_All_CSV_Files:
    CV_CSV_Data_List = []
    for file in os.listdir(data_Directory):
        if file.endswith(".csv"):
            CV_CSV_Data_List.append(file)
else:
    # Check to see if the Inputed CSV Files Exist
    for CV_CSV_Data in CV_CSV_Data_List:    
        if not os.path.isfile(data_Directory + CV_CSV_Data):
            print("The File ", data_Directory + CV_CSV_Data," Mentioned Does NOT Exist")
            exit()

# Create Output Folder if None
try:
    outputData = data_Directory +  "Full Time CV Curve Animation/"
    os.mkdir(outputData)
# Else, Continue On
except:
    pass

# ---------------------------------------------------------------------------#

def addPeakCurrent(addingDict, addingKey, addingValue):
    # see if the Key is in the Dictionary
    currentValues = addingDict.get(addingKey, [])
    # If Not, Add it
    if currentValues == []:
        addingDict[addingKey] = []
    # Add the Value to the Dictionary
    addingDict[addingKey].append(addingValue)
    return addingDict

def findPeakNum(peakDict, peakNum, potentialAdding):
    for oldPeakNum in peakDict.keys():
        peakPotentialAv = sum(peakDict[oldPeakNum])/len(peakDict[oldPeakNum])
        if potentialAdding < peakPotentialAv + peakError  and potentialAdding > peakPotentialAv - peakError:
            return oldPeakNum
    if peakDict.keys():
        return max(peakDict.keys()) + 1
    else:
        return 1
    

# -------------------- Extract and Plot the Ip Data -------------------------#

# For Eaqch CSV File, Extract the Important Data and Plot
for CV_CSV_Data in CV_CSV_Data_List: 
    
    # ----------------- Convert Data to Excel Format ------------------------#
    
    # Rename the File with an Excel Extension
    base = os.path.splitext(CV_CSV_Data)[0]
    excel_file = data_Directory + base + ".xlsx"
    # If the File is Not Already Converted: Convert
    if not os.path.isfile(excel_file):
        # Make Excel WorkBook
        wb = xl.Workbook()
        ws = wb.active
        # Write to Excel WorkBook
        with open(data_Directory + CV_CSV_Data) as f:
            reader = csv.reader(f, delimiter=',')
            for row in reader:
                ws.append(row)
        # Save as New Excel File
        wb.save(excel_file)
    elif skipIfDataAlreadyExcel:
        print("You already renamed the '.csv' to '.xlsx'; Hence, we are skipping file:", base)
        continue
    else:
        print("You already renamed the '.csv' to '.xlsx'; Will Redo Plots Anyways")
    
    # Load Data from New Excel File
    WB = xl.load_workbook(excel_file)
    WB_worksheets = WB.worksheets
    Main = WB_worksheets[0]
        
    # -----------------------------------------------------------------------#

    # ----------------------- Extract Run Info ------------------------------#
    
    # Set Initial Variables from last Run to Zero
    scanRate = None; sampleInterval = None; highVolt = None; lowVolt = None; dataIndex = None
    # Loop Through the Info Section and Extract the Needxed Run Info from Excel
    for cell in Main['A']:
        # Get Cell Value
        cellVal = cell.value
        if cellVal == None:
            continue
        
        # Find the Scan Rate (Volts/Second)
        if cellVal.startswith("Scan Rate (V/s) = "):
            scanRate = float(cellVal.split(" = ")[-1])
        # Find the Sample Interval (Voltage Different Between Points)
        elif cellVal.startswith("Sample Interval (V) = "):
            sampleInterval = float(cellVal.split(" = ")[-1])
        # Find the Highest Voltage
        elif cellVal.startswith("High E (V) = "):
            highVolt = float(cellVal.split(" = ")[-1])
        # Find the Lowest Voltage
        elif cellVal.startswith("Low E (V) = "):
            lowVolt = float(cellVal.split(" = ")[-1])
        elif cellVal == "Segment 1:":
            startSegment = cell.row
        elif cellVal == "Potential/V":
            dataIndex = cell.row + 2
            break
    # Find the X Axis Width
    xRange = (highVolt - lowVolt)*2
    # Find Point/Scan
    pointsPerScan = int(xRange/sampleInterval)
    # Adjust Which Cycle you Start at
    skipOffset = int(numInitCyclesToSkip*pointsPerScan)
    dataIndex += skipOffset
    # Total Frames (Will Round Down to Remove Incomplete Scans); Frame = Cycle = 2 Segments
    totalFrames = math.floor((Main.max_row - dataIndex + 1)/pointsPerScan)
    numberOfSegments = totalFrames*2
    
    # -----------------------------------------------------------------------#
  
    # ---------------- Get Data and Place into Frames -----------------------#
    
    # Get the Data
    current = []; potential = []; time = [0]
    currentFrames = []; potentialFrames = []; timeFrame = []
    lowCurrent = 10000; highCurrent = -10000
    for rowA, rowB in Main.iter_rows(min_col=1, min_row=dataIndex, max_col=2, max_row=Main.max_row):
        # Get Potential and Current Data Points
        potentialVal = rowA.value
        currentVal = rowB.value
        
        # If There is No More Data, Stop Recording
        if potentialVal == None:
            break
        
        # Add Data to Current Frame
        potential.append(float(potentialVal))
        current.append(float(currentVal))
        if len(potential) > 1:
            timeGap =  abs(potential[-1] - potential[-2]) / scanRate
            time.append(time[-1] + timeGap)
        
        # If Done Collecting Data, Collect as Frame and Start a New Frame
        if len(potential) >= pointsPerScan:
            # Add Current Frame
            potentialFrames.append(potential)
            currentFrames.append(current)
            timeFrame.append(time)
            # Keep Running Track of Min/Max Current (Y-Axis)
            lowCurrent = min(current + [lowCurrent])
            highCurrent = max(current + [highCurrent])
            # Reset for New Frame
            current = []; potential = []; time = [time[-1] + timeGap]
            
    # -----------------------------------------------------------------------#
    
    # ------------------- Find Ip Data and Store Info -----------------------#
    
    peakCurrent = {"peakCurrentForward":{}, "peakCurrentReverse":{}}
    cycleNumber = {"cycleNumberForward":{}, "cycleNumberReverse":{}}
    peakPotential = {"peakPotentialForward":{}, "peakPotentialReverse":{}}
    cycleNum = 0
    for rowA in Main.iter_rows(min_col=1, min_row=startSegment, max_col=1, max_row=dataIndex - 4 - skipOffset):
        cellVal = rowA[0].value
        if cellVal == None:
            continue
    
        # Skip Over Bad Segments
        if numInitCyclesToSkip > 0:
            if rowA[0].row == startSegment:
                continue
            if cellVal.startswith("Segment "):
                segment = float(cellVal[:-1].split("Segment ")[-1])
                numInitCyclesToSkip -= 0.5
            if numInitCyclesToSkip != 0:
                continue
        
        # Find the Current Segment
        if cellVal.startswith("Segment "): 
            # Extract Segment Info
            segment = float(cellVal[:-1].split("Segment ")[-1])
            peakNum = 1
            # It is a New Cycle Everytime we Scan Forwards
            if segment%2 == 1:
                cycleNum += 1
            # Stop if Next is Incomplete Segment; Else Continue Looping
            if cycleNum*2 == numberOfSegments+1:
                print("HERE")
                break
            continue
        elif cellVal.startswith("Ep = "):
            Ep = [float(x.split(":")[0]) for x in re.findall("-?\d+.?\d*(?:[Ee]-\d+)?", cellVal)][0]
        # find the Peak Current in the Segment
        elif cellVal.startswith("ip = "):
            Ip = [float(x.split(":")[0]) for x in re.findall("-?\d+.?\d*(?:[Ee]-\d+)?", cellVal)][0]
            # If it is an Odd Segment, it is the Forwards Peak
            if segment%2 == 1:
                peakNum = findPeakNum(peakPotential["peakPotentialForward"], peakNum, Ep)
                peakPotential["peakPotentialForward"] = addPeakCurrent(peakPotential["peakPotentialForward"], peakNum, Ep)
                peakCurrent["peakCurrentForward"] = addPeakCurrent(peakCurrent["peakCurrentForward"], peakNum, Ip)
                cycleNumber["cycleNumberForward"] = addPeakCurrent(cycleNumber["cycleNumberForward"], peakNum, cycleNum)
            else:
                peakNum = findPeakNum(peakPotential["peakPotentialReverse"], peakNum, Ep)
                peakPotential["peakPotentialReverse"] = addPeakCurrent(peakPotential["peakPotentialReverse"], peakNum, Ep)
                peakCurrent["peakCurrentReverse"] = addPeakCurrent(peakCurrent["peakCurrentReverse"], peakNum, Ip)
                cycleNumber["cycleNumberReverse"] = addPeakCurrent(cycleNumber["cycleNumberReverse"], peakNum, cycleNum)

    
    # ---------------------------------------------------------------------- #
  
    # --------------------- Plot and Save the Data --------------------------#
    
    # Initialize Plot Figure (Must be BEFORE MovieWriter Initialization)
    figWidth = 20
    figHeight = 8
    if (peakCurrent["peakCurrentForward"] != {} or peakCurrent["peakCurrentReverse"])and showPeakCurrent:
        if showFullInfo:
            fig, ax = plt.subplots(2, 2, sharey=False, sharex = False, figsize=(figWidth,figHeight))
            axLeft = ax[0,0]
            axRight = ax[0,1]
            axLowerLeft = ax[1,0]
            axLowerRight = ax[1,1]
        else:
            fig, ax = plt.subplots(1, 2, sharey=False, sharex = False, figsize=(figWidth,figHeight))
            axLeft = ax[0]
            axRight = ax[1]
    else:
        fig, axLeft = plt.subplots(1, 1, sharey=False, sharex = False, figsize=(figWidth/2,figHeight))
            
    # Initialize Movie Writer for Plots
    FFMpegWriter = manimation.writers['ffmpeg']
    metadata = dict(title=base, artist='Matplotlib', comment='Movie support!')
    writer = FFMpegWriter(fps=7, metadata=metadata)
    movieGraphLeftCurrent, = axLeft.plot([], [], 'b-', linewidth=1, alpha = 1)
    if seePastCVData:
        movieGraphLeftPrev, = axLeft.plot([], [], 'b-', linewidth=1, alpha = 0.1)
    
    # Set Axis X,Y Limits
    axLeft.set_xlim(lowVolt, highVolt)
    axLeft.set_ylim(lowCurrent, highCurrent)
    # Label Axis + Add Title
    axLeft.set_title("Time Dependant CV")
    axLeft.set_xlabel("Potential (Volts)")
    axLeft.set_ylabel("Current (Amps)")
    
    if (peakCurrent["peakCurrentForward"] != {} or peakCurrent["peakCurrentReverse"])and showPeakCurrent:
        # Repeat for Second Graph
        peakPlotHolder = {}
        for peakNum in peakCurrent["peakCurrentForward"].keys():
            peakPlotHolder["peakCurrentForward|" + str(peakNum)] = axRight.plot([], [], '-o', c=peakCurrentRightColorOrder["peakCurrentForward|" + str(peakNum)], linewidth=1)[0]
        for peakNum in peakCurrent["peakCurrentReverse"].keys():
            peakPlotHolder["peakCurrentReverse|" + str(peakNum)] = axRight.plot([], [], '-o', c=peakCurrentRightColorOrder["peakCurrentReverse|" + str(peakNum)], linewidth=1)[0]
        # Set Axis X,Y Limits: Taking min/max of both dictionarys and then min/max between the values
        if peakCurrent["peakCurrentForward"] == {}:
            peakLow = min((min(peakCurrent["peakCurrentReverse"].values(),key=min)))
            peakHigh = max((max(peakCurrent["peakCurrentReverse"].values(),key=max)))
        elif peakCurrent["peakCurrentReverse"] == {}:
            peakLow = min((min(peakCurrent["peakCurrentForward"].values(),key=min)))
            peakHigh = max((max(peakCurrent["peakCurrentForward"].values(),key=max)))
        else:
            peakLow = min(min(min(peakCurrent["peakCurrentForward"].values(),key=min)), min((min(peakCurrent["peakCurrentReverse"].values(),key=min))))
            peakHigh = max(max(max(peakCurrent["peakCurrentForward"].values(),key=max)), max((max(peakCurrent["peakCurrentReverse"].values(),key=max))))
        axRight.set_xlim(0, totalFrames)
        axRight.set_ylim(peakLow - 0.2*abs(peakLow), peakHigh + 0.2*abs(peakLow))
        # Label Axis + Add Title
        axRight.set_title("Peak Current Over CV Scan")
        axRight.set_xlabel("Cycle Number")
        axRight.set_ylabel("Peak Current (Amps)")
        
        # If Adding Potential and Standard Deviation
        if showFullInfo:
            # Repeat for Third and Forth Graph
            for peakNum in peakPotential["peakPotentialForward"].keys():
                peakPlotHolder["peakPotentialForward|" + str(peakNum)] = axLowerLeft.plot([], [], '-o', c=peakCurrentRightColorOrder["peakCurrentForward|" + str(peakNum)], linewidth=1)[0]
                peakPlotHolder["peakStatsForward|" + str(peakNum)] = axLowerRight.plot([], [], '-o', c=peakCurrentRightColorOrder["peakCurrentForward|" + str(peakNum)], linewidth=1)[0]
            for peakNum in peakPotential["peakPotentialReverse"].keys():
                peakPlotHolder["peakPotentialReverse|" + str(peakNum)] = axLowerLeft.plot([], [], '-o', c=peakCurrentRightColorOrder["peakCurrentReverse|" + str(peakNum)], linewidth=1)[0]
                peakPlotHolder["peakStatsReverse|" + str(peakNum)] = axLowerRight.plot([], [], '-o', c=peakCurrentRightColorOrder["peakCurrentReverse|" + str(peakNum)], linewidth=1)[0]
            # Set Axis X,Y Limits: Taking min/max of both dictionarys and then min/max between the values
            if peakPotential["peakPotentialForward"] == {}:
                peakLow = min((min(peakPotential["peakPotentialReverse"].values(),key=min)))
                peakHigh = max((max(peakPotential["peakPotentialReverse"].values(),key=max)))
            elif peakPotential["peakPotentialReverse"] == {}:
                peakLow = min((min(peakPotential["peakPotentialForward"].values(),key=min)))
                peakHigh = max((max(peakPotential["peakPotentialForward"].values(),key=max)))
            else:
                peakLow = min(min(min(peakPotential["peakPotentialForward"].values(),key=min)), min((min(peakPotential["peakPotentialReverse"].values(),key=min))))
                peakHigh = max(max(max(peakPotential["peakPotentialForward"].values(),key=max)), max((max(peakPotential["peakPotentialReverse"].values(),key=max))))
            axLowerRight.set_xlim(0, totalFrames)
            axLowerRight.set_ylim(0,30) # Greater Than 30 is Considered NOT Acceptable (Hence, I Will Cut it Off)
            axLowerLeft.set_xlim(0, totalFrames)
            axLowerLeft.set_ylim(peakLow - 0.2*abs(peakLow), peakHigh + 0.2*abs(peakLow))
            # Label Axis + Add Title
            axLowerLeft.set_title("Peak Potential Over CV Scan")
            axLowerLeft.set_xlabel("Cycle Number")
            axLowerLeft.set_ylabel("Peak Potential (Volts)")
            # Set Axis and Title for Stats
            axLowerRight.set_title("Peak Current's Coefficient of Variation Over CV Scan")
            axLowerRight.set_xlabel("Cycle Number")
            axLowerRight.set_ylabel("Standard Error")
            
    
    fig.tight_layout(pad=2.0)
    
    # Open Movie Writer and Add Data
    with writer.saving(fig, outputData + base + ".mp4", 300):
        # Add Frames in the Order for Showing
        for frameNum in range(totalFrames):
            # Set Left Side
            x = potentialFrames[frameNum]
            y = currentFrames[frameNum]
            t = timeFrame[frameNum]
            axLeft.legend(["RunTime = " + str(round(t[0],2)) + " Seconds"], loc="upper left")
            movieGraphLeftCurrent.set_data(x, y)
            if seePastCVData:
                movieGraphLeftPrev.set_data(potentialFrames[:frameNum], currentFrames[:frameNum])
                    
            # Set Right Side
            if (peakCurrent["peakCurrentForward"] != {} or peakCurrent["peakCurrentReverse"])and showPeakCurrent:
                legendListRight = []
                legendListLowerLeft = []
                legendListLowerRight = []
                for currentPeak in peakPlotHolder.keys():
                    # Catagorize the Peak
                    peakDirection, peakNum = currentPeak.split("|")
                    if peakDirection[:11] != "peakCurrent":
                        continue
                    EpDirection = peakDirection[:4] + 'Potential' + peakDirection[11:]
                    potentialPeak = EpDirection + "|" + peakNum
                    statsPeak =  peakDirection[:4] + 'Stats' + peakDirection[11:]  + "|" + peakNum
                    cyclePeak = "cycleNumber" + peakDirection[11:]
                    # If No Peak Current, Label NA
                    try:
                        indexFrame = cycleNumber[cyclePeak][int(peakNum)].index(frameNum+1)
                    except:
                        legendListRight.append(peakDirection[11:] + " Peak (" + peakNum + "): NA")
                        if showFullInfo:
                            legendListLowerLeft.append(peakDirection[11:] + " Peak (" + peakNum + "): NA")
                            legendListLowerRight.append(peakDirection[11:] + " Peak (" + peakNum + "): NA")
                        continue
                    # Get Peak Current, Potential, and Cycle
                    Ip = peakCurrent[peakDirection][int(peakNum)][:indexFrame+1]
                    Ep = peakPotential[EpDirection][int(peakNum)][:indexFrame+1]
                    cycle = cycleNumber[cyclePeak][int(peakNum)][:indexFrame+1]
                    # Get Statistics of Peak Current
                    IpSTD = [0]
                    CoefficientofVariationList = ["NA"]
                    CoefficientofVariation = "NA"
                    for IpNum in range(1,len(Ip)):
                        IpSTD.append(stat.stdev(Ip[:IpNum + 1]))
                        CoefficientofVariation = abs(IpSTD[-1]/stat.mean(Ip))*100
                        CoefficientofVariationList.append(CoefficientofVariation)
                    # Plot Peak Currents
                    movieGraphRight = peakPlotHolder[currentPeak]
                    legendListRight.append(peakDirection[11:] + " Peak (" + peakNum + "): Ep = " + "%.3g"%Ep[-1] + " Volts ; Ip = " + "%.4g"%Ip[-1] + " Amps")
                    movieGraphRight.set_data(cycle, Ip)
                    if showFullInfo:
                        # Plot Peak Potential
                        movieGraphLowerLeft = peakPlotHolder[potentialPeak]
                        legendListLowerLeft.append(peakDirection[11:] + " Peak (" + peakNum + "): Ep = " + "%.3g"%Ep[-1] + " Volts ; Ip = " + "%.4g"%Ip[-1] + " Amps")
                        movieGraphLowerLeft.set_data(cycle, Ep)
                        # ---------------------------------------- #
                        # Get Plot for Coefficient of VariationList
                        movieGraphLowerRight = peakPlotHolder[statsPeak]
                        # Add Legend: No Statistics Availible for One Data Point
                        if CoefficientofVariationList[-1] == "NA":
                            legendListLowerRight.append(peakDirection[11:] + " Peak (" + peakNum + "): NA")
                        else:
                            legendListLowerRight.append(peakDirection[11:] + " Peak (" + peakNum + "): Peak Current's Coefficient of Variation = " + "%.3g"%CoefficientofVariation + "%")
                        # Plot the Data
                        movieGraphLowerRight.set_data(cycle[1:], CoefficientofVariationList[1:])
                axRight.legend(legendListRight, loc="upper left")
                if showFullInfo:
                    axLowerLeft.legend(legendListLowerLeft, loc="upper left")
                    axLowerRight.legend(legendListLowerRight, loc="upper left")

            # Write to Video
            writer.grab_frame()
    # Close Writer
    #writer.close()
    plt.show()
        
    # -----------------------------------------------------------------------#

    