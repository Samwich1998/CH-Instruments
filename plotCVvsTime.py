
"""
Need to Install on the Anaconda Prompt:
    $ conda install openpyxl
    $ conda install seaborn
"""
import csv
import os
import openpyxl as xl
import matplotlib.pyplot as plt
import pandas as pd
%matplotlib qt 

# -------------------------- User Can Edit ----------------------------------#

use_All_CSV_Files = True # If False, Populate the CV_CSV_Data_List Yourself
data_Directory = "../NASA Project Cortisol/Prussian Blue/12-24-2020 Precursor Analysis/" # The Folder with the CSV Files

# ---------------------------------------------------------------------------#

# -------------------------- Find File Names --------------------------------#

if use_All_CSV_Files:
    CV_CSV_Data_List = []
    for file in os.listdir(data_Directory):
        if file.endswith(".csv"):
            CV_CSV_Data_List.append(file)
else:
    CV_CSV_Data_List = [
         'Fe3+ Ferricyanide Time 60 Min 0.25M NaOH.csv',
         ]
    
    # Check to see if the Inputed CSV Files Exist
    for CV_CSV_Data in CV_CSV_Data_List:    
        if not os.path.isfile(data_Directory + CV_CSV_Data):
            print("The File ", data_Directory + CV_CSV_Data," Mentioned Does NOT Exist")
            exit()
 
# Create Output Folder if None
try:
    outputData = data_Directory +  "Full Time CV Curve/"
    os.mkdir(outputData)
# Else, Continue On
except:
    pass

# ---------------------------------------------------------------------------#

# -------------------- Extract and Plot the Ip Data -------------------------#

# For Eaqch CSV File, Extract the Important Data and Plot
for figNum, CV_CSV_Data in enumerate(CV_CSV_Data_List):    
    
    # ----------------- Convert Data to Excel Format ------------------------#
    
    # Rename the File with an Excel Extension
    base = os.path.splitext(CV_CSV_Data)[0]
    excel_file = data_Directory + base + ".xlsx"
    # If the File is Not Already Converted: Convert
    if not os.path.isfile(excel_file):
        # Make Excel WorkBook
        wb = xl.Workbook()
        ws = wb.active
        # Write to Excel WorkBook
        with open(data_Directory + CV_CSV_Data) as f:
            reader = csv.reader(f, delimiter=',')
            for row in reader:
                ws.append(row)
        # Save as New Excel File
        wb.save(excel_file)
    else:
        print("You already renamed the '.csv' to '.xlsx'")
    
    # Load Data from New Excel File
    WB = xl.load_workbook(excel_file) 
    WB_worksheets = WB.worksheets
    Main = WB_worksheets[0]
        
    # -----------------------------------------------------------------------#
    
    # ------------------------ Find Data and Time ---------------------------#
    
    # Find Starting Point
    df = pd.read_excel(excel_file)
    colA = Main['A'][0].value
    dataIndex = df[df[colA]=='Potential/V'].index[0] + 4
    # FInd Time Gaps
    timeIndex = df[df[colA] == 'Scan Rate (V/s) = 0.1'].index[0] + 1
    scanRate = float(Main['A'][timeIndex].value.split(" = ")[-1]) # Volts/Sec
    
    current = []
    potential = []
    time = []
    for rowA, rowB in Main.iter_rows(min_col=1, min_row=dataIndex, max_col=2, max_row=Main.max_row):
        # Get Cell Value
        potentialVal = rowA.value
        currentVal = rowB.value
        
        # If There is No More Data, Stop Recording
        if potentialVal == None:
            break
        
        # Store Data to plot
        potential.append(float(potentialVal))
        current.append(float(currentVal))
        if len(time) == 0:
            time.append(0)
        else:
            timeGap =  abs(potential[-1] - potential[-2]) / scanRate
            time.append(time[-1] + timeGap)
    
    # -----------------------------------------------------------------------#
  
    # --------------------- Plot and Save the Data --------------------------#
    
    # Plot and Save
    fig = plt.figure(figNum)
    ax = fig.add_subplot(111, projection='3d')
    
    # Scatter Plot
    skipData = 40 # Set to 1 to Plot All Points; Too Crowded
    ax.scatter(time[::skipData], potential[::skipData], current[::skipData], "o", c=current[::skipData], cmap='viridis', linewidth=0.1, s=30)
    # Connected Plot
    #ax.plot_trisurf(time, potential, current, cmap='viridis', edgecolor='none')
    
    ax.set_title('CV Plot of ' + base);
    ax.set_xlabel("Time (Sec)")
    ax.set_ylabel("Voltage (V)")
    ax.set_zlabel("Current (A)")
    fig.tight_layout(pad=5)
    fig.savefig(outputData + base + ".png", dpi=300)
    plt.show() # Must be the Last Line
    
    # -----------------------------------------------------------------------#

    